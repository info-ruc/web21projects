import os
import re

import openpyxl
import pandas as pd
from openpyxl import load_workbook

from openpyxl.workbook import Workbook
from tqdm import *

img_re = re.compile(r'<img.*?>')

def clean_data(path):
    for doc_name in tqdm(os.listdir(path)):
        new_path = path + '/' + str(doc_name)
        # print(path)
        wb = openpyxl.load_workbook(new_path)
        ws = wb.active
        rows = ws.max_row
        delete_list = []
        for row in range(1, rows):
            comment = str(ws.cell(row=row, column=1).value)
            comment = str(bytes(comment, encoding='utf-8').decode('utf-8').encode('gbk', 'ignore').decode('gbk'))
            comment = re.sub(pattern=img_re, repl='', string=comment)
            if len(comment) == 0 or comment.isspace() or comment == 'None':
                delete_list.append(row)
            ws.cell(row=row, column=1).value = comment
        # delete rows
        delete_list.reverse()
        for row in delete_list:
            ws.delete_rows(row, 1)
        wb.save(new_path)


if __name__ == '__main__':
    path_negative = './dataset_tieba/negative'
    path_positive = './dataset_tieba/positive'

    clean_data(path_negative)
    clean_data(path_positive)
