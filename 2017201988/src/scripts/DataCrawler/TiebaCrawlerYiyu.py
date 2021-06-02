#encoding=utf-8

'''
百度贴吧主题列表的url示例
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E&pn=0
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E&ie=utf-8&pn=50
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E&ie=utf-8&pn=100

贴吧帖子的url示例
http://tieba.baidu.com/p/4965638364
http://tieba.baidu.com/p/4966830781

帖子html示例：
<div id="post_content_103484160275" class="d_post_content j_d_post_content ">            天天德州里面的十日签活动，拿福字换了奖励，却没用奖励，用十天的换来的福却用完了，我就呵呵了，TM<a href="http://jump.bdimg.com/safecheck/index?url=rN3wPs8te/pL4AOY0zAwhz3wi8AXlR5gsMEbyYdIw620q6P4KNTgxO2jJ4m3Rd0yt6XyibVFgk4fbLMgytUg5Z4Q4OO30Ri8SAtVjSx1NjJY8K44RtEayPWmgLpz8y36hTdCT9IUs4T05hvgchh9CzxCZVKzLsIy7/LrcfWaVD0dJnZ4aIMtppuQye2MG4bWMDxm7iZ2BjQ="  class="ps_cb" target="_blank" onclick="$.stats.track(0, 'nlp_ps_word',{obj_name:'腾讯'});$.stats.track('Pb_content_wordner','ps_callback_statics')">腾讯</a>就是这么玩人的吗？？</div>
<div id="post_content_103484761972" class="d_post_content j_d_post_content ">            我也没到账</div>

帖子分页html示例
	<li class="l_pager pager_theme_5 pb_list_pager"><span class="tP">1</span>
<a href="/p/4889568200?pn=2">2</a>
<a href="/p/4889568200?pn=3">3</a>
<a href="/p/4889568200?pn=4">4</a>
<a href="/p/4889568200?pn=2">下一页</a>
<a href="/p/4889568200?pn=4">尾页</a>
</li>

'''
import os
import time
import urllib
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import socket
import re
import requests
import json
import pandas as pd

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from tqdm import *

timeout = 10
content_re = re.compile(r"post_content_\d+")
tiezi_url_re = re.compile(r"/p/\d+")
tiezi_pn_re = re.compile(r"/p/\d+\?pn=(?P<pn>\d+)")

def get_html(url):
    try:
        r = requests.get(url,timeout=30)
        r.raise_for_status()
        #这里我们知道百度贴吧的编码是utf-8，所以手动设置的。爬去其他的页面时建议使用：
        # r.endcodding = r.apparent_endconding
        r.encoding='utf-8'
        return r.text
    except:
        return " ERROR "

def read_url(url):
    try:
        fid = urllib.request.urlopen(url)
        soup = BeautifulSoup(fid, "lxml")
    except:
        print("error: open %s"%url)
        return None, None
    return fid, soup

class Tiezi(object):
    base_url = "http://tieba.baidu.com"

    def __init__(self, tiezi_url, is_check_tiezi = 0, all_tiezi_address = []):
        self.tiezi_url_ = tiezi_url

        self.url_ = self.join_url()
        print(self.url_)

        self.pn_ = 1  # 帖子页码，默认一页
        self.fid_ = []  # 每页的内容
        self.soup_ = []  # 每页的soup
        self.url_list_ = []  # 每页的url

        fid, soup = read_url(self.url_)
        if not fid:
            return
        if is_check_tiezi and self.url_ in all_tiezi_address:
            print(self.url_, " is repeated!")
            return

        self.fid_.append(fid)
        self.soup_.append(soup)
        self.url_list_.append(self.url_)

        # 检查帖子是否多页
        pf = self.soup_[0].find_all("li", {"class": "l_pager pager_theme_5 pb_list_pager"})
        if len(pf) == 0:
            return
        pf_urls = pf[0].find_all("a")
        for pf_url in pf_urls:
            mo = tiezi_pn_re.search(str(pf_url))
            cur_pn = int(mo.groupdict()["pn"])
            self.pn_ = cur_pn if cur_pn > self.pn_ else self.pn_

        # 拼接每页的url，并打开
        for pn in range(2, self.pn_ + 1):
            cur_url = "%s?pn=%d" % (self.url_, pn)
            print(cur_url)
            fid, soup = read_url(cur_url)

            if not fid:
                continue

            self.fid_.append(fid)
            self.soup_.append(soup)
            self.url_list_.append(cur_url)


    def join_url(self):
        '''生成帖子的url'''
        return urljoin(Tiezi.base_url, self.tiezi_url_)

    def get_content(self):
        '''解析页面获得所需要的信息'''
        all_text = []
        for i, soup in enumerate(self.soup_):
            contents = soup.find_all('div', attrs={'class': 'l_post l_post_bright j_l_post clearfix'})
            # 通过循环找到每个帖子里的我们需要的信息：
            for content in contents:
                try:
                    data_field = json.loads(content['data-field'])
                    comment = {}

                    comment['user_id'] = data_field['author']['user_id']
                    comment['user_name'] = data_field['author']['user_name']
                    comment['portrait'] = data_field['author']['portrait']
                    comment['post_id'] = data_field['content']['post_id']
                    comment['content_text'] = data_field['content']['content']

                    div_post = content.find_all('div', attrs={'class' : {"post-tail-wrap"}})
                    for temp1 in div_post:
                        temp2 = temp1.find_all('span', attrs={'class': {'tail-info'}})
                        for temp3 in temp2:
                            if ('来自' in temp3.text.strip()):
                                continue
                            elif ('楼' in temp3.text.strip()):
                                comment['lou_num'] = temp3.text.strip()
                            else:
                                comment['post_time'] = temp3.text.strip()

                    all_text.append(comment)

                except:
                    print('出了点小问题')

        return all_text

class Zhuye(object):
    def __init__(self, keyword):
        self.url_ = "http://tieba.baidu.com/f?kw=%s"%(urllib.parse.quote(keyword))

    def get_page_url(self, pn=1):
        '''生成指定页的url'''
        return "%s&pn=%d"%(self.url_, 50*(pn-1))

    def get_tiezi_list(self, pn):
        url = self.get_page_url(pn)
        print(url)

        fid, soup = read_url(url)
        if not fid:
            return []

        tiezi_list = []

        links = soup.find_all("a")
        for link in links:
            if "href" in link.attrs and tiezi_url_re.match(link["href"]):
                tiezi_list.append(link["href"])

        return tiezi_list


if __name__ == '__main__':
    socket.setdefaulttimeout(timeout)

    # is_check_tiezi = 1

    keyword = '抑郁'
    ori_filename = ['raw_data_yiyu.xlsx', 'raw_data_yiyu_new.xlsx']
    filename = 'raw_data_yiyu_new2.xlsx'

    zhuye = Zhuye(keyword)

    raw_data = pd.DataFrame()
    for single_name in ori_filename:
        df = pd.DataFrame(pd.read_excel(single_name))
        raw_data = raw_data.append(df, ignore_index=True)
    all_post_id_ori = raw_data['post_id'].values
    #print(all_post_id_ori)

    for i in range(1, 210): # 爬取的页数，每爬一整页的帖子储存一整页
        print("现在开始爬取第", i, "页！")
        tiezi_list = zhuye.get_tiezi_list(i)

        all_post_id_new = []
        wb = Workbook()
        if (os.path.isfile(filename)):
            wb = load_workbook(filename=filename)
            all_post_id_new = pd.DataFrame(pd.read_excel(filename))['post_id'].values
            # all_tiezi_address = pd.DataFrame(pd.read_excel(filename))['url'].values
        ws = wb.worksheets[0]
        row = ws.max_row
        if row == 1:
            ws.cell(row=row, column=1).value = 'user_id'
            ws.cell(row=row, column=2).value = 'user_name'
            ws.cell(row=row, column=3).value = 'portrait'
            ws.cell(row=row, column=4).value = 'post_id'
            ws.cell(row=row, column=5).value = 'content_text'
            ws.cell(row=row, column=6).value = 'lou_num'
            ws.cell(row=row, column=7).value = 'post_time'
            ws.cell(row=row, column=8).value = 'url'
            ws.cell(row=row, column=9).value = 'crawl_time'
            row += 1

        need_update = 0
        with tqdm(total=len(tiezi_list)) as pbar:
            for url in tiezi_list:
                # tiezi = Tiezi(url, is_check_tiezi, all_tiezi_address)
                tiezi = Tiezi(url)
                contents = tiezi.get_content()
                if len(contents) == 0:
                    print("the length of contents is 0!")
                for content in contents:
                    if content['post_id'] in all_post_id_ori:
                        continue
                    if content['post_id'] in all_post_id_new:
                        continue
                    if len(content['content_text']) == 0:
                        continue
                    ws.cell(row=row, column=1).value = content['user_id']
                    ws.cell(row=row, column=2).value = content['user_name']
                    ws.cell(row=row, column=3).value = content['portrait']
                    ws.cell(row=row, column=4).value = content['post_id']
                    ws.cell(row=row, column=5).value = content['content_text']
                    ws.cell(row=row, column=6).value = content['lou_num']
                    ws.cell(row=row, column=7).value = content['post_time']
                    ws.cell(row=row, column=8).value = 'http://tieba.baidu.com' + url
                    ws.cell(row=row, column=9).value = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                    row +=1
                    need_update = 1
                pbar.update(1)
        wb.save(filename)
        print("第", i, "页爬取完毕！\n")
        if need_update == 0:
            print("本页暂无帖子更新，终止爬取！")
            break
