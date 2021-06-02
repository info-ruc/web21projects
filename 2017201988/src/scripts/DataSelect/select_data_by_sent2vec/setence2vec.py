import jieba
import sent2vec
from openpyxl import Workbook
from tqdm import tqdm

model = sent2vec.Sent2vecModel()
model.load_model('my_model.bin')

sent_count = 0
all_sentences = []

with open("train_data.txt", 'r') as ori_f:
    lines = ori_f.readlines()
    for line in lines:
        sent_count+=1

with tqdm(total=sent_count) as pbar:
    with open("train_data.txt", 'r') as ori_f:
        wb = Workbook()
        ws = wb.worksheets[0]
        row = 1

        lines = ori_f.readlines()
        for line in lines:
            if line in all_sentences:
                pbar.update(1)
                continue

            seg_list = jieba.cut(line, cut_all=False)
            line_cut = " ".join(seg_list)

            emb = model.embed_sentence(line_cut)
            ws.cell(row=row, column=1).value = line
            ws.cell(row=row, column=2).value = line_cut
            ws.cell(row=row, column=3).value = str(emb[0])
            row += 1

            all_sentences.append(line)
            pbar.update(1)

        wb.save("train_data_vector.xlsx")
