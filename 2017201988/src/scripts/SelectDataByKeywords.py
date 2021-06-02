import os
import pandas as pd
from openpyxl import load_workbook

from openpyxl.workbook import Workbook
from tqdm import *


if __name__ == '__main__':

    raw_filename = 'raw_data_yiyu.xlsx'
    store_filename = 'selected_data_yiyu.xlsx'
    raw_data = pd.DataFrame(pd.read_excel(raw_filename))
    store_data = pd.DataFrame(pd.read_excel(store_filename))
    all_selected_post_id = store_data['post_id'].values
    wb = load_workbook(filename=store_filename)
    ws = wb.worksheets[0]
    store_row = ws.max_row

    key = ['确诊', '医院', '吃药', '轻度抑郁', '中度抑郁', '重度抑郁', '患病', '治疗',
           '发作']

    with tqdm(total=raw_data.shape[0]) as pbar:
        for index, row in raw_data.iterrows():
            useful_key = ''
            for single_key in key:
                if single_key in str(row['content_text']):
                    if row['post_id'] not in all_selected_post_id:
                        print("通过关键词【", single_key, '】筛选出', row['user_id'])
                        useful_key += single_key
                        useful_key += ' '
            if len(useful_key) > 0:
                ws.cell(row=store_row, column=1).value = row['user_id']
                ws.cell(row=store_row, column=2).value = row['user_name']
                ws.cell(row=store_row, column=3).value = row['portrait']
                ws.cell(row=store_row, column=4).value = row['post_id']
                ws.cell(row=store_row, column=5).value = useful_key
                ws.cell(row=store_row, column=6).value = row['content_text']
                ws.cell(row=store_row, column=7).value = row['lou_num']
                ws.cell(row=store_row, column=8).value = row['post_time']
                ws.cell(row=store_row, column=9).value = row['url']
                ws.cell(row=store_row, column=10).value = row['crawl_time']
                store_row += 1
            pbar.update(1)

    wb.save(store_filename)