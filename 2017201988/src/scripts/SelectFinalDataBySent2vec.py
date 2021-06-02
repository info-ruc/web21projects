import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from tqdm import *

img_re_jpg = re.compile(r'<img.*?class="BDE_[a-zA-Z]+".*?src="[^"]*\.jpg".*?>')
img_re_png = re.compile(r'<img.*?class="BDE_[a-zA-Z]+".*?src="[^"]*\.png".*?>')

def split_sent(raw_sent):
    one_line = str(raw_sent)
    final_str = []
    for br_str in one_line.split('<br>'):
        if len(br_str) == 0:
            continue
        for br_dou_ju_str in re.split("[。，]+", br_str):
            final_single_str = br_dou_ju_str
            img_list_jpg = re.findall(img_re_jpg, final_single_str)
            if len(img_list_jpg) > 0:
                for rid_str in img_list_jpg:
                    final_single_str = final_single_str.replace(rid_str, '')
            img_list_png = re.findall(img_re_png, final_single_str)
            if len(img_list_png) > 0:
                for rid_str in img_list_png:
                    final_single_str = final_single_str.replace(rid_str, '')
            if len(final_single_str) < 2:
                continue
            final_str.append(final_single_str)
    return final_str

if __name__ == '__main__':
    raw_data = pd.DataFrame()
    # database_name = ["raw_data_yiyu.xlsx", "raw_data_yiyu_new.xlsx", "raw_data_yiyuzheng.xlsx",
    #                  "raw_data_yiyuzheng_new.xlsx"]
    database_name = ["raw_data_test.xlsx"]
    for single_name in database_name:
        df = pd.DataFrame(pd.read_excel(single_name))
        raw_data = raw_data.append(df, ignore_index=True)

    key_sent_database_name = "./select_data_by_sent2vec/critical_sentence_vector.xlsx"
    key_sent_data = pd.DataFrame(pd.read_excel(key_sent_database_name))

    store_filename = 'final_selected_data.xlsx'
    store_data = pd.DataFrame(pd.read_excel(store_filename))
    all_selected_user_id = store_data['user_id'].values
    wb = load_workbook(filename=store_filename)
    ws = wb.worksheets[0]
    store_row = ws.max_row + 1

    current_scan = 0

    with tqdm(total=raw_data.shape[0]) as pbar:
        new_selected_user_id = []
        for index_raw, row_raw in raw_data.iterrows():
            all_split_sent = split_sent(row_raw['content_text'])

            for index_sent, row_sent in key_sent_data.iterrows():
                if_correct = str(row_sent['if_correct'])
                if len(if_correct) != 0 and if_correct != 'nan':
                    if float(if_correct) == 0:
                        continue

                if_confirmed = 0
                for single_sent in all_split_sent:
                    print(single_sent)
                    print(row_sent['sentence'])
                    if single_sent == str(row_sent['sentence']):
                        exit()
                        print(row_sent['sentence'])
                        if_confirmed = 1
                        break

                if if_confirmed == 1:
                    if row_raw['user_id'] in all_selected_user_id:
                        continue
                    if row_raw['user_id'] in new_selected_user_id:
                        continue
                    new_selected_user_id.append(row_raw['user_id'])
                    ws.cell(row=store_row, column=1).value = row_raw['user_id']
                    ws.cell(row=store_row, column=2).value = row_raw['user_name']
                    ws.cell(row=store_row, column=3).value = row_raw['portrait']
                    ws.cell(row=store_row, column=4).value = row_raw['post_id']
                    ws.cell(row=store_row, column=5).value = row_sent['sentence']
                    ws.cell(row=store_row, column=6).value = row_raw['content_text']
                    ws.cell(row=store_row, column=7).value = row_raw['lou_num']
                    ws.cell(row=store_row, column=8).value = row_raw['post_time']
                    ws.cell(row=store_row, column=9).value = row_raw['url']
                    ws.cell(row=store_row, column=10).value = row_raw['crawl_time']
                    store_row += 1
                    break
            pbar.update(1)
            current_scan += 1
            if current_scan%5000 == 4999:
                wb.save(store_filename)

    # wb.save(store_filename)
