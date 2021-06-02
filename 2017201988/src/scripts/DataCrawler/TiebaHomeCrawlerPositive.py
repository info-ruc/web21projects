#encoding=utf-8

'''
百度贴吧主题列表的url示例
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E&pn=0
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E&ie=utf-8&pn=50
http://tieba.baidu.com/f?kw=%E5%A4%A9%E5%A4%A9%E5%BE%B7%E5%B7%9E&ie=utf-8&pn=100

贴吧帖子的url示例
http://tieba.baidu.com/p/4965638364
http://tieba.baidu.com/p/4966830781

贴吧用户主页url示例
http://tieba.baidu.com/home/main?id=tb.1.d2e190d1.sgJKxkfSntMGYl2U-X7rSA?t=1437833218

帖子html示例：
<div id="post_content_103484160275" class="d_post_content j_d_post_content ">            天天德州里面的十日签活动，拿福字换了奖励，却没用奖励，用十天的换来的福却用完了，我就呵呵了，TM<a href="http://jump.bdimg.com/safecheck/index?url=rN3wPs8te/pL4AOY0zAwhz3wi8AXlR5gsMEbyYdIw620q6P4KNTgxO2jJ4m3Rd0yt6XyibVFgk4fbLMgytUg5Z4Q4OO30Ri8SAtVjSx1NjJY8K44RtEayPWmgLpz8y36hTdCT9IUs4T05hvgchh9CzxCZVKzLsIy7/LrcfWaVD0dJnZ4aIMtppuQye2MG4bWMDxm7iZ2BjQ="  class="ps_cb" target="_blank" onclick="$.stats.track(0, 'nlp_ps_word',{obj_name:'腾讯'});$.stats.track('Pb_content_wordner','ps_callback_statics')">腾讯</a>就是这么玩人的吗？？</div>
<div id="post_content_103484761972" class="d_post_content j_d_post_content ">            我也没到账</div>

帖子分页html示例
<li class="l_pager pager_theme_5 pb_list_pager"><span class="tP">1</span>
<a href="/p/4889568200?pn=2">2</a>
<a href="/p/4889568200?pn=3">3</a>
<a href="/p/4889568200?pn=4">4</a>
<a href="/p/4889568200?pn=2">下一页</a>
<a href="/p/4889568200?pn=4">尾页</a>
</li>


'''
import os
import urllib
from datetime import datetime, timedelta

from bs4 import BeautifulSoup
from urllib.parse import urljoin
import itertools
import socket
import re
import requests
import json
import pandas as pd

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from tqdm import *


timeout = 10
content_re = re.compile(r"post_content_\d+")
teiba_re = re.compile(r"/f\?kw=[0-9A-Z%]+")
tiezi_url_re = re.compile(r"/p/\d+")
tiezi_pn_re = re.compile(r"/p/\d+\?see_lz=1&amp;pn=(?P<pn>\d+)")
img_re = re.compile(r'<img.*?class="BDE_[a-zA-Z]+".*?src="[^"]*\.jpg".*?>')


def mkdir(path):
    folder = os.path.exists(path)
    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径



def get_html(url):
    try:
        r = requests.get(url,timeout=30)
        r.raise_for_status()
        #这里我们知道百度贴吧的编码是utf-8，所以手动设置的。爬去其他的页面时建议使用：
        # r.endcodding = r.apparent_endconding
        r.encoding='utf-8'
        return r.text
    except:
        return " ERROR "

def read_url(url):
    try:
        fid = urllib.request.urlopen(url)
        soup = BeautifulSoup(fid, "lxml")
    except:
        print("error: open %s"%url)
        return None, None
    return fid, soup

class Tiezi(object):
    base_url = "http://tieba.baidu.com"
    def __init__(self, tiezi_url):
        self.tiezi_url_ = tiezi_url

        self.url_ = self.join_url()
        print(self.url_)

        self.pn_ = 1    # 帖子页码，默认一页
        self.fid_ = []  # 每页的内容
        self.soup_ = [] # 每页的soup
        self.url_list_ = [] # 每页的url
        
        fid, soup = read_url(self.url_)
        if not fid:
            return

        self.fid_.append(fid)
        self.soup_.append(soup)
        self.url_list_.append(self.url_)

        # 检查帖子是否多页
        pf = self.soup_[0].find_all("li", {"class":"l_pager pager_theme_5 pb_list_pager"})
        if len(pf) == 0:
            return
        pf_urls = pf[0].find_all("a")
        for pf_url in pf_urls:
            mo = tiezi_pn_re.search(str(pf_url))
            cur_pn = int(mo.groupdict()["pn"])
            self.pn_ = cur_pn if cur_pn > self.pn_ else self.pn_

        # 拼接每页的url，并打开
        for pn in range(2, self.pn_+1):
            cur_url = "%s&pn=%d"%(self.url_, pn)
            print(cur_url)
            fid, soup = read_url(cur_url)

            if not fid:
                continue
            
            self.fid_.append(fid)
            self.soup_.append(soup)
            self.url_list_.append(cur_url)


    def join_url(self):
        '''生成帖子的url'''
        return urljoin(Tiezi.base_url, self.tiezi_url_)

    def get_content(self):
        '''解析页面获得所需要的信息'''
        all_text = []
        for i, soup in enumerate(self.soup_):
            # print(soup)
            contents1 = soup.find_all('div', attrs={'class': 'l_post j_l_post l_post_bright noborder'})
            contents2 = soup.find_all('div', attrs={'class': 'l_post j_l_post l_post_bright'})
            contents3 = soup.find_all('div', attrs={'class': 'l_post j_l_post l_post_bright noborder_bottom'})

            if (len(contents1) > 0):
                for content1 in contents1:
                    content1_result = self.get_single_content(content1)
                    if (type(content1_result) != None.__class__):
                        all_text.append(content1_result)
            if (len(contents2) > 0):
                for content2 in contents2:
                    content2_result = self.get_single_content(content2)
                    if (type(content2_result) != None.__class__):
                        all_text.append(content2_result)
            if (len(contents3) > 0):
                for content3 in contents3:
                    content3_result = self.get_single_content(content3)
                    if (type(content3_result) != None.__class__):
                        all_text.append(content3_result)

        return all_text


    def get_single_content(self, content):
        try:
            # print(content)
            data_field = json.loads(content['data-field'])
            # print(data_field)
            comment = {}

            comment['user_id'] = data_field['author']['user_id']
            comment['user_name'] = data_field['author']['user_name']
            comment['portrait'] = data_field['author']['portrait']
            comment['post_id'] = data_field['content']['post_id']
            div_content = content.find_all('div', attrs={'class': {"d_post_content j_d_post_content clearfix"}})
            img_list = re.findall(img_re, str(div_content[0]))

            comment['content_text'] = ''
            content_text = div_content[0].find_all(text=True)
            count = 0
            for single_str in content_text:
                if count > 0:
                    comment['content_text'] += '<br>'
                comment['content_text'] += str(single_str).strip()
                if len(comment['content_text']) > 0:
                    count += 1

            if len(img_list) > 0:
                for str_img in img_list:
                    comment['content_text'] = comment['content_text'] + str(str_img)

            comment['lou_num'] = data_field['content']['post_no']
            comment['post_time'] = data_field['content']['date']

            return comment
        except:
            print('出了点小问题')

class Home(object):
    def __init__(self, keyword):
        self.url_ = "http://tieba.baidu.com/home/main?id="+(urllib.parse.quote(keyword)+'#') # '#'代表展示所有帖子

    def get_tiezi_list(self):

        fid, soup = read_url(self.url_)
        if not fid:
            return []

        tiezi_list = []
        tiezi_type = []

        post = soup.find_all('div', attrs={'thread_name'})
        for single_post in post:
            links = single_post.find_all("a")
            is_kaixin = 0
            is_kuailedabenying = 0

            for link in links:
                if "href" in link.attrs and teiba_re.match(link["href"]): # 匹配贴子所在贴吧
                    if "开心" in link['title']:
                        is_kaixin = 1
                    if "快乐大本营" in link['title']:
                        is_kaixin = 1

            if is_kaixin == 1:
                for link in links:
                    if "href" in link.attrs and tiezi_url_re.match(link["href"]):
                        tiezi_list.append(link["href"].split('?')[0]+'?see_lz=1')
                        tiezi_type.append('开心吧')

            if is_kuailedabenying == 1:
                for link in links:
                    if "href" in link.attrs and tiezi_url_re.match(link["href"]):
                        tiezi_list.append(link["href"].split('?')[0]+'?see_lz=1')
                        tiezi_type.append('快乐大本营吧')


        return zip(tiezi_list, tiezi_type)


if __name__ == '__main__':
    socket.setdefaulttimeout(timeout)

    filename = 'raw_data_kuailedabenying.xlsx'
    selected_data = pd.DataFrame(pd.read_excel(filename))
    # print(selected_data)
    path = "./positive/"
    crawl_corrected = []

    database = pd.DataFrame()
    database_name = ["raw_data_kuailedabenying.xlsx"]
    # database_name = ["raw_data_yiyu_new.xlsx"]
    for single_name in database_name:
        df = pd.DataFrame(pd.read_excel(single_name))
        database = database.append(df, ignore_index=True)

    selected_data = selected_data.drop_duplicates("user_id")

    for index, row in selected_data.iterrows():
        # crawl_corrected.append(index)
        # if row['is_crawled'] == 1:
        #     print(str(row['user_id']) + ' is already handled')
        #     continue

        portrait = row['portrait']
        # portrait = 'tb.1.fa9803c9.ESjDJXor8VgPbG-uuoqXpA'
        user_id = row['user_id']
        base_time_late = datetime.strptime(row['post_time'], '%Y-%m-%d %H:%M')
        base_time_former = base_time_late - timedelta(days=180)

        filename = str(user_id) + '.xlsx'

        # mkdir(path + str(user_id))
        wb = Workbook()
        ws = wb.worksheets[0]
        row = ws.max_row
        if row == 1:
            ws.cell(row=row, column=1).value = 'content_text'
            ws.cell(row=row, column=2).value = 'lou_num'
            ws.cell(row=row, column=3).value = 'post_time'
            ws.cell(row=row, column=4).value = 'url'
            row += 1

        tiezi_info = list(zip(*Home(portrait).get_tiezi_list()))
        tiezi_list = []
        tiezi_type = []
        if (len(tiezi_info) != 0):
            tiezi_list = tiezi_info[0]
            tiezi_type = tiezi_info[1]

        crawled_post_id = []
        # crawl data from homepage
        with tqdm(total=len(tiezi_list)) as pbar:
            tiezi_count = 0
            for url in tiezi_list:
                tiezi = Tiezi(url)
                contents = tiezi.get_content()
                for content in contents:
                    if len(contents) == 0:
                        print("the length of contents is 0!")
                    post_time = datetime.strptime(content['post_time'], '%Y-%m-%d %H:%M')
                    if post_time < base_time_former or post_time > base_time_late:
                        continue
                    ws.cell(row=row, column=1).value = content['content_text']
                    ws.cell(row=row, column=2).value = content['lou_num']
                    ws.cell(row=row, column=3).value = content['post_time']
                    ws.cell(row=row, column=4).value = 'https://tieba.baidu.com' + url
                    row +=1
                    crawled_post_id.append(content['post_id'])
                pbar.update(1)
                tiezi_count += 1

        # get data from crawled database
        stored_data = database[database['user_id'].isin([user_id])]
        for index_s, row_s in stored_data.iterrows():
            post_time = datetime.strptime(row_s['post_time'], '%Y-%m-%d %H:%M')
            if post_time < base_time_former or post_time > base_time_late:
                continue
            if row_s['post_id'] in crawled_post_id:
                continue
            ws.cell(row=row, column=1).value = row_s['content_text']
            ws.cell(row=row, column=2).value = row_s['lou_num']
            ws.cell(row=row, column=3).value = row_s['post_time']
            ws.cell(row=row, column=4).value = row_s['url']
            row += 1

        wb.save(path+'/'+filename)

    # for i in crawl_corrected:
    #     selected_data.set_value(i, 'is_crawled', 1)

    # pd.DataFrame(selected_data).to_excel('selected_data.xlsx', index=False, header=True)
