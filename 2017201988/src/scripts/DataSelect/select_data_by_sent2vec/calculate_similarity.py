import random
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from scipy import spatial


def get_vector(ori_vector):
    raw_vector = list(filter(None, str(ori_vector[1:-1]).split(' ')))
    cleaned_vector = []
    for i in range(len(raw_vector)):
        cleaned_vector.append(float(raw_vector[i].replace('\n', '')))
    return np.array(cleaned_vector)


if __name__ == '__main__':
    raw_filename = 'train_data_vector.xlsx'
    store_filename = 'critical_sentence_vector.xlsx'
    raw_data = pd.DataFrame(pd.read_excel(raw_filename))
    store_data = pd.DataFrame(pd.read_excel(store_filename))
    ori_row_num = len(store_data)

    all_selected_sentences = store_data['sentence'].values
    wb = load_workbook(filename=store_filename)
    ws = wb.worksheets[0]
    store_row = ws.max_row + 1

    with tqdm(total=raw_data.shape[0]) as pbar:
        for index, row in raw_data.iterrows():
            if row['sentence'] in all_selected_sentences:
                pbar.update(1)
                continue

            vector_now = get_vector(row['vector'])

            similarity_list = []
            selected_row_list = random.sample(range(ori_row_num), min(ori_row_num,5))
            for selected_row in selected_row_list:
                if_correct = str(store_data.iloc[selected_row]['if_correct'])
                if len(if_correct) != 0 and if_correct != 'nan':
                    if float(if_correct) == 0:
                        continue
                base_vector = get_vector(store_data.iloc[selected_row]['vector'])
                single_similarity = 1 - spatial.distance.cosine(base_vector, vector_now)
                similarity_list.append(single_similarity)

            ave_similarity = np.mean(similarity_list)
            if_have_deny_word = 0
            deny_list = ['没去', '想去', '咨询', '女友', '男友', '老公', '老婆', '吗', '女朋友', '女票',
                         '妈妈', '母亲', '爸爸', '父亲', '他', '她', '请问', '儿子', '女儿', '男朋友', '男票',
                         '？', '?', '女盆友', '男盆友', '孩子', '先生', '是不是', '老妈', '没有', '弟弟',
                         '怎么', '什么', '闺蜜']
            for deby_word in deny_list:
                if deby_word in str(row['sentence']):
                    if_have_deny_word = 1
                    break

            if (ave_similarity > 0.90 and if_have_deny_word == 0):
                ws.cell(row=store_row, column=2).value = row['sentence']
                ws.cell(row=store_row, column=3).value = row['sentence_cut']
                ws.cell(row=store_row, column=4).value = row['vector']
                ws.cell(row=store_row, column=5).value = ave_similarity
                store_row += 1
                print(row['sentence'])
                print(ave_similarity)
                print()

            pbar.update(1)

    # wb.save(store_filename)